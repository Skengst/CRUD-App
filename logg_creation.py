import csv
import mysql.connector
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from db_config import database_host,database_user, database_password, database_name, database_port
def excel_transfer(login_count):
    try:
        wb = load_workbook("login history.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Year", "Month", "Day", "Hour", "Count"])
    
    current_time = datetime.now()
    hour=current_time.hour
    day=current_time.day
    month=current_time.month
    year=current_time.year

    ws.append([year,month,day,hour,login_count])
    wb.save("login history.xlsx")

def insert_data():
    csv_file = "user_login.csv"

    connection = mysql.connector.connect(
    host=database_host,
    user=database_user,
    password=database_password,
    database=database_name,
    port=database_port,
    )
    
    cursor = connection.cursor()
    login_count = 0
    try:
        with open(csv_file, "r") as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                username, time = row
                cursor.execute(f"INSERT INTO login_history (username, time) VALUES (%s, %s)", (username, time))
                login_count +=1
        connection.commit()
        excel_transfer(login_count)

        with open("user_login.csv", mode="w") as file:
            pass
    except Exception as e:
        print(f"Error: {e}")
    finally:
        cursor.close()
        connection.close()

def etl_process():
    while True:
        insert_data()
        time.sleep (3600)

if __name__ == "__main__":
    print ("Running ETL process")
    etl_process()
